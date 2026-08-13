#!/usr/bin/env python3
"""xls2spread — 把「固定資產建設改良擴充計畫及成本效益分析表」類的 Excel 轉成可手動調列高的對開表 HTML。

用法:  python3 xls2spread.py 表.xls [-o out.html] [-s 工作表名]
輸出的 HTML 直接用瀏覽器開，調完列高／行距後 Cmd+P 存 PDF。
"""
import argparse, html, json, os, shutil, subprocess, sys, tempfile

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

EMU_IN = 914400
BW = {'hair': 0.25, 'thin': 0.5, 'medium': 1.0, 'thick': 1.5,
      'double': 1.0, 'dashed': 0.5, 'dotted': 0.5, 'mediumDashed': 1.0}
FONT_STACK = {
    '微軟正黑體': "'Microsoft JhengHei','PingFang TC','Noto Sans CJK TC',sans-serif",
    '新細明體': "'PMingLiU','Songti TC','Noto Serif CJK TC',serif",
    '細明體': "'MingLiU','Songti TC',serif",
    '華康中明體': "'DFKai-SB','Songti TC',serif",
    '華康中黑體': "'PingFang TC','Microsoft JhengHei',sans-serif",
    'Times New Roman': "'Times New Roman',Times,serif",
}


def to_xlsx(path):
    if path.lower().endswith(('.xlsx', '.xlsm')):
        return path
    soffice = shutil.which('soffice') or '/Applications/LibreOffice.app/Contents/MacOS/soffice'
    tmp = tempfile.mkdtemp()
    subprocess.run([soffice, '--headless', '--convert-to', 'xlsx', '--outdir', tmp, path],
                   check=True, capture_output=True)
    return os.path.join(tmp, os.path.splitext(os.path.basename(path))[0] + '.xlsx')


def fmt_num(v, nf):
    """夠用的數值格式：會計格式的 0 顯示 '-'，其餘看小數位與千分位。"""
    if isinstance(v, bool):
        return str(v)
    secs = nf.split(';')
    if v == 0 and len(secs) >= 3 and '-' in secs[2]:
        return '-'
    if nf in ('General', '@', ''):
        return f'{v:g}'
    dec = 2 if '0.00' in nf else 0
    return f'{v:,.{dec}f}' if '#,##0' in nf else f'{v:.{dec}f}'


def color_of(c):
    if c is None or c.type != 'rgb' or not c.rgb or not isinstance(c.rgb, str):
        return None
    rgb = c.rgb[-6:]
    return None if rgb == '000000' else '#' + rgb


def runs_html(v):
    """rich text -> HTML。原檔是用白色文字當「換行後的縮排」，
    所以白字一定是行首：在它前面補換行，換行點才不會隨字型寬度跑掉。"""
    if not isinstance(v, openpyxl.cell.rich_text.CellRichText):
        return html.escape(str(v))
    out, plain = [], ''
    for r in v:
        t = getattr(r, 'text', r)
        f = getattr(r, 'font', None)
        col = color_of(f.color) if f is not None and f.color is not None else None
        if col in ('#FFFFFF', '#ffffff') and plain and not plain.endswith('\n'):
            out.append('\n')
            plain += '\n'
        esc = html.escape(t)
        out.append(f'<span style="color:{col}">{esc}</span>' if col else esc)
        plain += t
    return ''.join(out)


class Styles:
    """把用到的儲存格樣式收斂成 CSS class。"""

    def __init__(self):
        self.map, self.css = {}, []

    def cls(self, cell, nohz=False):
        """nohz: 資料列不要橫的列線，只留直的欄線"""
        f, a, b = cell.font, cell.alignment, cell.border
        sig = (f.name, f.sz, bool(f.b), color_of(f.color), a.horizontal, a.vertical,
               bool(a.wrap_text), a.indent or 0,
               b.left.style, b.right.style,
               None if nohz else b.top.style, None if nohz else b.bottom.style)
        if sig in self.map:
            return self.map[sig]
        name = f'c{len(self.map)}'
        self.map[sig] = name
        d = [f'font-family:{FONT_STACK.get(sig[0], "sans-serif")}',
             f'font-size:calc({sig[1] or 11}pt * var(--fs) * var(--k))']
        if sig[2]:
            d.append('font-weight:700')
        if sig[3]:
            d.append(f'color:{sig[3]}')
        # Excel 的「分散對齊」= 字撐滿整格
        d.append('text-align:justify;text-align-last:justify' if sig[4] == 'distributed'
                 else f'text-align:{sig[4] or "left"}')
        if sig[7]:
            d.append(f'padding-left:calc({sig[7] * 0.5}em + 2px)')
        d.append('white-space:' + ('pre-wrap' if sig[6] else 'pre'))
        for side, st in zip(('left', 'right', 'top', 'bottom'), sig[8:]):
            if st:
                d.append(f'border-{side}:{BW.get(st, 0.5)}pt solid #000')
        self.css.append(f'.{name}{{{";".join(d)}}}')
        return name


VAL = {'center': 'vc', 'bottom': 'vb'}


def build(path, sheet=None):
    xlsx = to_xlsx(path)
    wbv = openpyxl.load_workbook(xlsx, data_only=True)
    wbs = openpyxl.load_workbook(xlsx, rich_text=True)
    ws = wbv[sheet] if sheet else next(w for w in wbv.worksheets if w.print_area)
    wt = wbs[ws.title]

    pa = ws.print_area[0] if isinstance(ws.print_area, list) else ws.print_area
    c1, r1, c2, r2 = range_boundaries(pa.split('!')[-1].replace('$', ''))
    split = ([b.id for b in ws.col_breaks.brk] or [c2])[0]           # 左頁 = c1..split
    halves = {'l': list(range(c1, split + 1)), 'r': list(range(split + 1, c2 + 1))}

    tr = ws.print_title_rows                                          # e.g. '$1:$6'
    hdr_last = int(tr.split(':')[1].strip('$')) if tr else r1 - 1
    hdr_rows = list(range(r1, hdr_last + 1))
    data_rows = list(range(hdr_last + 1, r2 + 1))

    merged = {}
    skip = set()
    for rng in ws.merged_cells.ranges:
        mc1, mr1, mc2, mr2 = rng.min_col, rng.min_row, rng.max_col, rng.max_row
        merged[(mr1, mc1)] = (mr2 - mr1 + 1, mc2 - mc1 + 1)
        for r in range(mr1, mr2 + 1):
            for c in range(mc1, mc2 + 1):
                if (r, c) != (mr1, mc1):
                    skip.add((r, c))

    S = Styles()

    def cell(r, c, nohz=False):
        if (r, c) in skip:
            return None
        cv, cs = ws.cell(r, c), wt.cell(r, c)
        v = cv.value
        if v is None:
            txt = ''
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            txt = html.escape(fmt_num(v, cs.number_format))
        else:
            txt = runs_html(cs.value)
        rs, csn = merged.get((r, c), (1, 1))
        o = {'t': txt, 'c': S.cls(cs, nohz)}
        va = VAL.get(cs.alignment.vertical)
        if va:
            o['v'] = va
        if rs > 1:
            o['rs'] = rs
        if csn > 1:
            o['cs'] = csn
        return o

    def framed(r, cols):
        """有垂直框線才算表格本體；只有底線（天線）的仍屬標題區。"""
        return any(getattr(ws.cell(r, c).border, s).style for s in ('left', 'right') for c in cols)

    allc = halves['l'] + halves['r']
    title_rows = [r for r in hdr_rows if not framed(r, allc)]
    head_rows = [r for r in hdr_rows if r not in title_rows]

    # 合計列（下框線 medium）與其後的附註列 -> 釘在天地線位置
    pinned = [r for r in data_rows if any(ws.cell(r, c).border.bottom.style == 'medium' for c in allc)]
    foot_from = pinned[-1] if pinned else r2 + 1
    body_rows = [r for r in data_rows if r < foot_from]
    foot_rows = [r for r in data_rows if r >= foot_from]

    def rowh(r):
        rd = ws.row_dimensions.get(r)
        return round(rd.height if rd and rd.height else (ws.sheet_format.defaultRowHeight or 16.5), 2)

    def pack(rows, nohz=False):
        return [{'h': rowh(r),
                 'l': [c for c in (cell(r, i, nohz) for i in halves['l']) if c is not None],
                 'r': [c for c in (cell(r, i, nohz) for i in halves['r']) if c is not None]} for r in rows]

    # <col min max width> 可能一筆蓋好幾欄，要自己展開；直接查 column_dimensions[字母] 會拿到錯的值
    cw = {}
    for cd in ws.column_dimensions.values():
        if cd.width:
            for c in range(cd.min or 0, (cd.max or 0) + 1):
                cw[c] = cd.width

    def widths(side):
        w = [cw.get(c, 8.43) for c in halves[side]]
        t = sum(w)
        return [round(x / t * 100, 4) for x in w]

    brk = {b.id for b in ws.row_breaks.brk}
    m, ph = ws.page_margins, 297 if ws.page_setup.orientation != 'landscape' else 210
    pw = 210 if ws.page_setup.orientation != 'landscape' else 297

    # rowBreaks 的 id 是「這一列之後分頁」，所以下一頁從 id+1 開始
    breaks = sorted(body_rows.index(r) + 1 for r in brk if r in body_rows and r != body_rows[-1])
    # Excel fitToPage 會整體縮放列印，縮放率沒存在檔裡；用「最滿的一頁剛好塞進版心」回推
    hs = [rowh(r) for r in body_rows]
    pgsum, acc = [], 0.0
    for i, h in enumerate(hs):
        if i in breaks and i:
            pgsum.append(acc)
            acc = 0.0
        acc += h
    pgsum.append(acc + sum(rowh(r) for r in foot_rows))
    fixed = sum(rowh(r) for r in title_rows) + sum(rowh(r) for r in head_rows)
    body_pt = (ph - m.top * 25.4 - m.bottom * 25.4) * 72 / 25.4
    k = min(1.0, round((body_pt - 2) / (fixed + max(pgsum)), 4)) if pgsum else 1.0

    return {
        'k': k,
        'name': os.path.basename(path),
        'sheet': ws.title,
        'page': {'w': pw, 'h': ph, 'ml': round(m.left * 25.4, 2), 'mr': round(m.right * 25.4, 2),
                 'mt': round(m.top * 25.4, 2), 'mb': round(m.bottom * 25.4, 2)},
        'cols': {'l': widths('l'), 'r': widths('r')},
        'title': pack(title_rows),
        'head': pack(head_rows),
        'rows': pack(body_rows, nohz=True),
        'foot': pack(foot_rows),
        'breaks': breaks,
        'css': S.css,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xls')
    ap.add_argument('-o', '--out')
    ap.add_argument('-s', '--sheet')
    a = ap.parse_args()
    data = build(a.xls, a.sheet)
    tpl = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'template.html'), encoding='utf-8').read()
    out = a.out or os.path.splitext(a.xls)[0] + '.html'
    tpl = (tpl.replace('/*__CSS__*/', '\n'.join(data.pop('css')))
              .replace('__PW__', str(data['page']['w'])).replace('__PH__', str(data['page']['h']))
              .replace('"__DATA__"', json.dumps(data, ensure_ascii=False)))
    open(out, 'w', encoding='utf-8').write(tpl)
    print(out)


if __name__ == '__main__':
    main()
